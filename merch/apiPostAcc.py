# 
# Python auto _POST data through Api 
# Using Request _POST <php> to insert Data Merch to DB Server
#...


#   ---------- NOTICE: ----------------
#   DATA IN EXCEL MUST BE NOT NONE|| BLANK
# 
# 
# 

import requests
import openpyxl

def get_dataFromExcelByRow(sheet,row_start):

    acc_merch = {
    'acc_id': '',
    'user': '',
    'alias': '',
    'tier': '',
    'alive_status': '',

    }

    #.. c_price must be NOT NONE || BLANK
    c_acc_id = sheet.cell(row=row_start,column=2).value

    #.. c_productype must be NOT NONE || BLANK
    c_user = sheet.cell(row=row_start,column=3).value

    #.. c_color must be NOT NONE || BLANK
    c_alias = sheet.cell(row=row_start,column=4).value

    #.. c_brandname must be NOT NONE || BLANK
    c_tier = sheet.cell(row=row_start,column=5).value

    #.. c_url_image must be NOT NONE || BLANK
    c_alive_status = sheet.cell(row=row_start,column=6).value

    # 
    #   MUST BE CHECK NULL - NONE DATA GET FROM EXCEL
    #       ...............
    #

    print("      ")
    print(".....Data at Row:............." + str(row_start))
    print("      ")

    if c_acc_id is not None:
        acc_merch['acc_id'] = c_acc_id
    else:
        print("Can not Insert into DB due to c_acc_id is NONE||EMPTY")
    

    if c_user is not None:
        acc_merch['user'] = c_user
    else:
        print("Can not Insert into DB due to c_user is NONE||EMPTY")

    if c_alias is not None:
        acc_merch['alias'] = c_alias
    else:
        print("Can not Insert into DB due to c_alias is NONE||EMPTY")
    
    if c_tier is not None:
        acc_merch['tier'] = c_tier
    else:
        print("Can not Insert into DB due to c_tier is NONE||EMPTY")

    if c_alive_status is not None:
        acc_merch['alive_status'] = c_alive_status
    else:
        print("Can not Insert into DB due to c_alive_status is NONE||EMPTY")

    

    print("..........................................................................................")


    return acc_merch

#------------------------ --------- ------------------------------
#------------------------ MAIN METHOD ------------------------------


Data_Temp = "./accmerch/accmerch.xlsx"
f_Data_Temp = openpyxl.load_workbook(Data_Temp,data_only=True)
sheet_Temp = f_Data_Temp['FinalData']

row_start = 2
limit_Row = sheet_Temp.max_row

while row_start <= limit_Row:

    # Get Data from Excel Row by Row
    # Return Json Type 
    #..
    data = get_dataFromExcelByRow(sheet_Temp,row_start)

    print(data)

    # Request post to Push Data to API-PHP
    response = requests.post('http://localhost/controlpro/merch/api_post_acc_merch.php', data=data)

    # Check the response status
    if response.status_code == 200:
        print("Data inserted successfully!")

        c_up_Server = sheet_Temp.cell(row=row_start,column=17)
        c_up_Server.value = 1

        #....
        #   Change File to save data new
        f_Data_Temp.save(Data_Temp)


    else:
        print("Error inserting data:", response.text)

    row_start = row_start + 1

    


